# 9)B) Demonstrate python program to read the data from the spreadsheet and 
# write the data in to the spreadsheet

import openpyxl
workbook = openpyxl.load_workbook('example.xlsx')
worksheet = workbook.active
value = worksheet.cell(row=1, column=1).value
print(f"Value of cell A1: {value}")
worksheet.cell(row=2, column=1, value="Hello, world!")
workbook.save('example.xlsx')

# import openpyxl

# wb = openpyxl.load_workbook('example.xlsx')
# ws = wb.active

# print(f"Value of cell A1: {ws['A1'].value}")
# ws['A2'] = "Hello, world!"

# wb.save('example.xlsx')
